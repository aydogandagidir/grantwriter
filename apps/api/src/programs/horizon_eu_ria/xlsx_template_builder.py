"""Build the Horizon Europe Lump Sum budget XLSX scaffold.

Per docs/07 §3.5: EC's official Lump Sum template has 12 sheets and
embedded macros (XLSM) — we don't reproduce the macros. The committed
binary at ``templates/lump_sum_budget_2026.xlsx`` is a minimal
3-sheet scaffold (Partners / WP Costs / Summary) that the runtime
:func:`HorizonEURIAModule.export_xlsx_budget` fills with proposal data.

Replacing this with the real EC template is a separate task (Bluedev
internal — needs the latest binary from the EC portal). When that lands,
the ``export_xlsx_budget`` cell-write coordinates may need adjustment;
the contract surface (3 sheets, named columns) is preserved on purpose.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Sheet name conventions — load-bearing because export_xlsx_budget
# looks them up by name. Don't rename without updating the export.
SHEET_PARTNERS = "Partners"
SHEET_WP_COSTS = "WP Costs"
SHEET_SUMMARY = "Summary"

PARTNER_COLUMNS = ["Partner Name", "Country", "Entity Type"]
WP_COLUMNS = ["WP #", "WP Title", "Lead Partner", "Person Months", "Total (EUR)"]
SUMMARY_ROWS = [
    ("Programme", "Horizon Europe RIA / IA"),
    ("Lump Sum Total (EUR)", ""),
    ("Number of Partners", ""),
    ("Number of WPs", ""),
]

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_LABEL_FONT = Font(bold=True)


def build_template() -> Workbook:
    """Build the 3-sheet Lump Sum scaffold and return the Workbook."""

    wb = openpyxl.Workbook()
    # The default sheet becomes Summary; rename + repurpose.
    summary = wb.active
    if summary is None:
        summary = wb.create_sheet(SHEET_SUMMARY)
    summary.title = SHEET_SUMMARY
    _build_summary_sheet(summary)

    partners = wb.create_sheet(SHEET_PARTNERS)
    _build_header_row(partners, PARTNER_COLUMNS)

    wp_costs = wb.create_sheet(SHEET_WP_COSTS)
    _build_header_row(wp_costs, WP_COLUMNS)

    return wb


def to_bytes() -> bytes:
    """Convenience: build the template and return its bytes."""

    buf = BytesIO()
    build_template().save(buf)
    return buf.getvalue()


def main(target: Path | None = None) -> Path:
    """CLI entrypoint: write the template to its committed path."""

    target = target or (
        Path(__file__).resolve().parent / "templates" / "lump_sum_budget_2026.xlsx"
    )
    target.parent.mkdir(parents=True, exist_ok=True)
    build_template().save(str(target))
    return target


# ── Helpers ────────────────────────────────────────────────────────────


def _build_header_row(ws: Worksheet, columns: list[str]) -> None:
    """Write a styled header row (bold, blue fill, white text)."""

    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    # Set a generous column width so the headers are legible without
    # auto-fit (which openpyxl doesn't compute).
    for col_idx, _ in enumerate(columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22


def _build_summary_sheet(ws: Worksheet) -> None:
    """Two-column key/value summary."""

    for row_idx, (label, value) in enumerate(SUMMARY_ROWS, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = _LABEL_FONT
        ws.cell(row=row_idx, column=2, value=value)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30


# Suppress mypy's "Any" complaint — `openpyxl.utils` doesn't ship type stubs.
_ = Any


__all__ = [
    "PARTNER_COLUMNS",
    "SHEET_PARTNERS",
    "SHEET_SUMMARY",
    "SHEET_WP_COSTS",
    "SUMMARY_ROWS",
    "WP_COLUMNS",
    "build_template",
    "main",
    "to_bytes",
]


if __name__ == "__main__":
    written = main()
    print(f"wrote: {written}")
