"""HE Lump Sum XLSX export tests.

Two layers:
1. Pure-function tests on ``HorizonEURIAModule.export_xlsx_budget`` —
   round-trip the bytes through openpyxl and assert cell content.
2. Celery-task tests on ``_render_and_upload_xlsx`` and
   ``generate_proposal_xlsx_task`` — fake storage, no Celery broker.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import uuid4

import openpyxl
import pytest
from src.programs import get_module
from src.programs.horizon_eu_ria.xlsx_template_builder import (
    SHEET_PARTNERS,
    SHEET_SUMMARY,
    SHEET_WP_COSTS,
    build_template,
    to_bytes,
)
from src.storage.supabase_storage import UploadResult
from src.tasks.exports import (
    XLSX_MIME,
    _render_and_upload_xlsx,
    _xlsx_storage_path,
)

# ── Template builder ────────────────────────────────────────────────────


def test_template_builder_creates_three_sheets() -> None:
    wb = build_template()
    assert SHEET_PARTNERS in wb.sheetnames
    assert SHEET_WP_COSTS in wb.sheetnames
    assert SHEET_SUMMARY in wb.sheetnames


def test_template_builder_to_bytes_round_trips() -> None:
    blob = to_bytes()
    wb = openpyxl.load_workbook(BytesIO(blob))
    assert SHEET_PARTNERS in wb.sheetnames
    # Header row populated.
    partners_header = [
        wb[SHEET_PARTNERS].cell(row=1, column=col).value for col in (1, 2, 3)
    ]
    assert partners_header == ["Partner Name", "Country", "Entity Type"]


# ── export_xlsx_budget ─────────────────────────────────────────────────


def _proposal_with_full_budget() -> dict[str, Any]:
    return {
        "id": str(uuid4()),
        "tenant_id": str(uuid4()),
        "programme_id": "horizon_eu_ria",
        "title": "GREENMOBILITY",
        "budget": {
            "total_eur": 4_500_000,
            "by_partner": [
                {"name": "ACME", "country": "NL", "entity_type": "SME"},
                {"name": "Beta Research", "country": "DE", "entity_type": "RTO"},
                {"name": "Gamma", "country": "TR", "entity_type": "SME"},
            ],
            "by_wp": [
                {
                    "wp_number": 1,
                    "title": "Requirements",
                    "lead_partner": "ACME",
                    "person_months": 12,
                    "total": 800_000,
                },
                {
                    "wp_number": 2,
                    "title": "Implementation",
                    "lead_partner": "Beta Research",
                    "person_months": 36,
                    "total": 2_400_000,
                },
            ],
        },
    }


def test_export_xlsx_budget_returns_bytes_with_partners_filled() -> None:
    module = get_module("horizon_eu_ria")
    blob = module.export_xlsx_budget(_proposal_with_full_budget())

    assert blob is not None
    assert len(blob) > 1000  # non-trivial workbook

    wb = openpyxl.load_workbook(BytesIO(blob))
    ws = wb[SHEET_PARTNERS]

    # Row 1 = header. Row 2-4 = partners.
    assert ws.cell(row=2, column=1).value == "ACME"
    assert ws.cell(row=2, column=2).value == "NL"
    assert ws.cell(row=2, column=3).value == "SME"
    assert ws.cell(row=3, column=1).value == "Beta Research"
    assert ws.cell(row=4, column=1).value == "Gamma"


def test_export_xlsx_budget_fills_wp_costs() -> None:
    module = get_module("horizon_eu_ria")
    blob = module.export_xlsx_budget(_proposal_with_full_budget())
    assert blob is not None

    wb = openpyxl.load_workbook(BytesIO(blob))
    ws = wb[SHEET_WP_COSTS]

    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Requirements"
    assert ws.cell(row=2, column=5).value == 800_000
    assert ws.cell(row=3, column=2).value == "Implementation"


def test_export_xlsx_budget_updates_summary() -> None:
    module = get_module("horizon_eu_ria")
    blob = module.export_xlsx_budget(_proposal_with_full_budget())
    assert blob is not None

    wb = openpyxl.load_workbook(BytesIO(blob))
    ws = wb[SHEET_SUMMARY]

    summary = {row[0].value: row[1].value for row in ws.iter_rows(values_only=False)}
    assert summary["Lump Sum Total (EUR)"] == 4_500_000
    assert summary["Number of Partners"] == 3
    assert summary["Number of WPs"] == 2


def test_export_xlsx_budget_returns_none_when_budget_missing() -> None:
    """No budget dict → no XLSX. The Saga skips the upload step."""

    module = get_module("horizon_eu_ria")
    assert module.export_xlsx_budget({"programme_id": "horizon_eu_ria"}) is None


def test_export_xlsx_budget_handles_empty_partners_and_wps() -> None:
    """Budget present but empty lists — XLSX renders with the template
    headers only (no data rows)."""

    module = get_module("horizon_eu_ria")
    proposal = {
        "id": str(uuid4()),
        "tenant_id": str(uuid4()),
        "programme_id": "horizon_eu_ria",
        "budget": {"total_eur": 0, "by_partner": [], "by_wp": []},
    }
    blob = module.export_xlsx_budget(proposal)
    assert blob is not None

    wb = openpyxl.load_workbook(BytesIO(blob))
    ws_partners = wb[SHEET_PARTNERS]
    # Row 2 should be empty (no data row past the header).
    assert ws_partners.cell(row=2, column=1).value is None


# ── Storage path helper ────────────────────────────────────────────────


def test_xlsx_storage_path_format() -> None:
    path = _xlsx_storage_path(tenant_id="tenant-x", proposal_id="prop-y")
    assert path.startswith("tenant/tenant-x/proposal/prop-y/budget-")
    assert path.endswith(".xlsx")


# ── _render_and_upload_xlsx ────────────────────────────────────────────


class _FakeStorage:
    """In-memory storage stub for upload tests."""

    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []

    async def store_and_sign(
        self,
        *,
        path: str,
        data: bytes,
        content_type: str,
        expires_in: int = 3600,
    ) -> UploadResult:
        self.calls.append(
            {
                "path": path,
                "size_bytes": len(data),
                "content_type": content_type,
                "expires_in": expires_in,
            }
        )
        return UploadResult(
            bucket="exports",
            path=path,
            signed_url=f"https://fake.signed/{path}",
            expires_in=expires_in,
        )


async def test_render_and_upload_xlsx_uploads_with_xlsx_mime() -> None:
    storage = _FakeStorage()
    proposal = _proposal_with_full_budget()

    result = await _render_and_upload_xlsx(proposal=proposal, storage=storage)

    assert result is not None
    assert result.bucket == "exports"
    assert len(storage.calls) == 1
    call = storage.calls[0]
    assert call["content_type"] == XLSX_MIME
    assert call["path"].endswith(".xlsx")


async def test_render_and_upload_xlsx_returns_none_for_no_xlsx_programme() -> None:
    """TÜBİTAK 1501 returns None from export_xlsx_budget — task skips upload."""

    storage = _FakeStorage()
    proposal = {
        "id": str(uuid4()),
        "tenant_id": str(uuid4()),
        "programme_id": "tubitak_1501",
        "budget": {"by_category": {}},
    }
    result = await _render_and_upload_xlsx(proposal=proposal, storage=storage)
    assert result is None
    assert storage.calls == []


async def test_render_and_upload_xlsx_requires_tenant_id() -> None:
    storage = _FakeStorage()
    proposal = _proposal_with_full_budget()
    proposal["tenant_id"] = ""  # missing
    with pytest.raises(ValueError, match="tenant_id"):
        await _render_and_upload_xlsx(proposal=proposal, storage=storage)
